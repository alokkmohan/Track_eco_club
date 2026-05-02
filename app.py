import streamlit as st
import pandas as pd
import io
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl

st.set_page_config(
    page_title="ECO Club - Notification Upload Status",
    page_icon="🌿",
    layout="wide"
)

DISTRICTS = {
    "Lakhimpur Kheri":  ("KHERI",       "KHERI",          True),
    "Auraiya":          ("AURAIYA",      "AURAIYA",        True),
    "Chandauli":        ("CHANDAULI",    "CHANDAULI",      True),
    "Bulandshahr":      ("BULANDSHAHR",  "BULANDSHAHR",    True),
    "Hathras":          ("HATHRAS",      "HATHRAS",        True),
    "Shravasti":        ("SHRAV|SHRAW",  "SHRAV|SHRAW",    False),
    "Gonda":            ("GONDA",        "GONDA",          True),
    "Basti":            ("BASTI",        "BASTI",          True),
    "Bareilly":         ("BAREILLY",     "BAREILLY",       True),
    "Ballia":           ("BALLIA",       "BALLIA",         True),
    "Rampur":           ("RAMPUR",       "RAMPUR",         True),
    "Moradabad":        ("MORADABAD",    "MORADABAD",      True),
    "Etawah":           ("ETAWAH",       "ETAWAH",         True),
    "Azamgarh":         ("AZAMGARH",     "AZAMGARH",       True),
    "Amethi":           ("AMETHI",       "AMETHI",         False),
    "Kannauj":          ("KANNAUJ",      "KANNAUJ",        True),
    "Jaunpur":          ("JAUNPUR",      "JAUNPUR",        True),
    "Hardoi":           ("HARDOI",       "HARDOI",         True),
    "Hapur":            ("HAPUR",        "HAPUR",          False),
    "Amroha":           ("AMROHA",       "AMROHA|JYOTIBA", False),
    "Bhadohi":          ("BHADOHI",      "BHADOI",         False),
    "Prayagraj":        ("PRAYAGRAJ",    "PRAYAGRAJ",      True),
    "Sambhal":          ("SAMBHAL",      "SAMBHAL",        False),
    "Mirzapur":         ("MIRZAPUR",     "MIRZAPUR",       True),
    "Meerut":           ("MEERUT",       "MEERUT",         True),
    "Farrukhabad":      ("FARRUKHABAD",  "FARRUKHABAD",    True),
}

def normalize_udise(val):
    s = str(val).strip().split('.')[0].lstrip('0')
    return s.zfill(10)

@st.cache_data
def load_secondary_list():
    return {
        sheet: pd.read_excel("data/Final_Secondary_School_List.xlsx", sheet_name=sheet)
        for sheet in ["Govt Schools", "Aided Schools", "Private Schools"]
    }

@st.cache_data
def load_notifications():
    return pd.read_excel("data/All_Schools_with_Notifications_UP.xlsx")

def has_private_data(secondary_data, dist_label):
    sec_pattern, _, exact = DISTRICTS[dist_label]
    priv_df  = secondary_data["Private Schools"]
    dist_col = [c for c in priv_df.columns if 'district' in c.lower()][0]
    if exact:
        rows = priv_df[priv_df[dist_col].astype(str).str.strip().str.upper() == sec_pattern]
    else:
        rows = priv_df[priv_df[dist_col].astype(str).str.strip().str.upper().str.contains(sec_pattern, regex=True, na=False)]
    return len(rows) > 0

def get_district_data(notifications_df, secondary_data, dist_label, extra_private_df=None):
    sec_pattern, notif_pattern, exact = DISTRICTS[dist_label]

    if exact:
        notif_d = notifications_df[notifications_df['District'].str.strip().str.upper() == notif_pattern].copy()
    else:
        notif_d = notifications_df[notifications_df['District'].str.strip().str.upper().str.contains(notif_pattern, regex=True, na=False)].copy()

    notif_d['UDISE_norm'] = notif_d['UDISE ID'].apply(normalize_udise)
    notif_set = set(notif_d['UDISE_norm'])

    result = {}
    for sheet, df in secondary_data.items():
        if sheet == "Private Schools" and extra_private_df is not None:
            d = extra_private_df.copy()
            dist_col  = [c for c in d.columns if 'district' in c.lower()][0]
            udise_col = [c for c in d.columns if 'udise' in c.lower()][0]
            d[dist_col] = dist_label
        else:
            dist_col  = [c for c in df.columns if 'district' in c.lower()][0]
            udise_col = [c for c in df.columns if 'udise' in c.lower()][0]
            if exact:
                d = df[df[dist_col].astype(str).str.strip().str.upper() == sec_pattern].copy()
            else:
                d = df[df[dist_col].astype(str).str.strip().str.upper().str.contains(sec_pattern, regex=True, na=False)].copy()
            d[dist_col] = dist_label

        d['UDISE_norm'] = d[udise_col].apply(normalize_udise)
        d[udise_col]    = d['UDISE_norm']
        d['Upload Status'] = d['UDISE_norm'].apply(lambda x: 'Uploaded' if x in notif_set else 'Pending')
        d = d.drop(columns=['UDISE_norm'])
        result[sheet] = d

    return result

def build_excel(district_data):
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    for sheet, df in district_data.items():
        df_out = df.reset_index(drop=True)
        df_out.index = df_out.index + 1
        df_out.index.name = 'S.No.'
        df_out.to_excel(writer, sheet_name=sheet, index=True)

    writer.close()
    output.seek(0)

    wb = openpyxl.load_workbook(output)
    green_fill  = PatternFill('solid', fgColor='C6EFCE')
    red_fill    = PatternFill('solid', fgColor='FFC7CE')
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))

    for ws in wb.worksheets:
        status_col    = None
        udise_col_idx = None
        for cell in ws[1]:
            if cell.value == 'Upload Status':
                status_col = cell.column
            if cell.value and 'udise' in str(cell.value).lower():
                udise_col_idx = cell.column
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin
                cell.alignment = Alignment(vertical='center')
                if udise_col_idx and cell.column == udise_col_idx:
                    cell.number_format = '@'
                    if cell.value:
                        cell.value = str(cell.value).zfill(10)
            if status_col:
                sc = ws.cell(row=row[0].row, column=status_col)
                if sc.value == 'Uploaded':
                    sc.fill = green_fill
                    sc.font = Font(bold=True, color='276221')
                elif sc.value == 'Pending':
                    sc.fill = red_fill
                    sc.font = Font(bold=True, color='9C0006')
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 45)
        ws.freeze_panes = 'B2'

    final = io.BytesIO()
    wb.save(final)
    final.seek(0)
    return final

def private_template_excel(dist_label):
    df = pd.DataFrame(columns=['District Name', 'Block Name', 'School Name', 'UDISE Code', 'Board', 'Managed By'])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Private Schools')
    output.seek(0)
    return output

def show_district_report(notifications_df, secondary_data, dist_label, extra_private=None):
    district_data = get_district_data(notifications_df, secondary_data, dist_label, extra_private)

    c1, c2, c3 = st.columns(3)
    for col_ui, sheet_name, emoji in zip(
        [c1, c2, c3],
        ["Govt Schools", "Aided Schools", "Private Schools"],
        ["🏛️", "🤝", "🏫"]
    ):
        df = district_data[sheet_name]
        total    = len(df)
        uploaded = (df['Upload Status'] == 'Uploaded').sum()
        pending  = total - uploaded
        pct      = round(uploaded / total * 100) if total > 0 else 0
        with col_ui:
            st.metric(
                label=f"{emoji} {sheet_name.replace(' Schools','')}",
                value=f"{uploaded} / {total}",
                delta=f"{pending} pending" if pending > 0 else "✅ All uploaded",
                delta_color="inverse"
            )
            st.progress(pct / 100)

    excel_data = build_excel(district_data)
    st.download_button(
        label=f"⬇️ {dist_label} — Full Report Download",
        data=excel_data,
        file_name=f"{dist_label}_ECO_Club_Status.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True
    )

    st.markdown("#### 📋 School-wise Detail")
    tabs = st.tabs(["🏛️ Govt Schools", "🤝 Aided Schools", "🏫 Private Schools"])
    for tab, sheet_name in zip(tabs, ["Govt Schools", "Aided Schools", "Private Schools"]):
        with tab:
            df = district_data[sheet_name].reset_index(drop=True)
            df.index = df.index + 1
            pending_df  = df[df['Upload Status'] == 'Pending']
            uploaded_df = df[df['Upload Status'] == 'Uploaded']
            sub1, sub2 = st.tabs([
                f"❌ Pending ({len(pending_df)})",
                f"✅ Uploaded ({len(uploaded_df)})"
            ])
            with sub1:
                if len(pending_df) > 0:
                    st.dataframe(pending_df, use_container_width=True)
                else:
                    st.success("Sab schools ne upload kar di! 🎉")
            with sub2:
                st.dataframe(uploaded_df, use_container_width=True)

# ─────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────

st.title("🌿 ECO Club — Notification Upload Status")
st.markdown("---")

if 'uploaded_private' not in st.session_state:
    st.session_state.uploaded_private = {}

with st.spinner("Data load ho raha hai..."):
    notifications_df = load_notifications()
    secondary_data   = load_secondary_list()

# Categorize districts
ready_districts   = []
pending_districts = []
for dist in sorted(DISTRICTS.keys()):
    if has_private_data(secondary_data, dist) or dist in st.session_state.uploaded_private:
        ready_districts.append(dist)
    else:
        pending_districts.append(dist)

col_left, col_right = st.columns([6, 4], gap="large")

# ── LEFT: Report ready districts ──
with col_left:
    st.markdown(f"### ✅ Report Ready Districts ({len(ready_districts)})")
    selected = st.selectbox(
        "Apna jila chunein",
        ["-- Jila chunein --"] + ready_districts,
        key="ready_sel"
    )
    if selected != "-- Jila chunein --":
        extra = st.session_state.uploaded_private.get(selected)
        with st.spinner(f"{selected} ki report ban rahi hai..."):
            show_district_report(notifications_df, secondary_data, selected, extra)

# ── RIGHT: Pending districts (private list missing) ──
with col_right:
    st.markdown(f"### ⚠️ Private List Pending ({len(pending_districts)})")

    if not pending_districts:
        st.success("Sabhi districts ki private school list available hai! 🎉")
    else:
        pending_sel = st.selectbox(
            "Jila chunein",
            ["-- Jila chunein --"] + pending_districts,
            key="pending_sel"
        )

        if pending_sel != "-- Jila chunein --":
            st.info(f"**{pending_sel}** ki private school list abhi upload nahi hui।")

            template = private_template_excel(pending_sel)
            st.download_button(
                label="📥 Template Download Karen",
                data=template,
                file_name=f"{pending_sel}_Private_School_Template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

            st.markdown("**Template bharke yahan upload karein:**")
            uploaded_list = st.file_uploader(
                "Private School List (.xlsx / .csv)",
                type=["xlsx", "csv"],
                key=f"priv_{pending_sel}"
            )

            if uploaded_list:
                try:
                    if uploaded_list.name.endswith('.csv'):
                        df_priv = pd.read_csv(uploaded_list)
                    else:
                        df_priv = pd.read_excel(uploaded_list)

                    udise_cols = [c for c in df_priv.columns if 'udise' in c.lower()]
                    if not udise_cols:
                        st.error("❌ File mein UDISE column nahi mila। Template use karein।")
                    else:
                        st.session_state.uploaded_private[pending_sel] = df_priv
                        st.success(f"✅ {pending_sel} ki list upload ho gayi! Left panel mein district aa jaega।")
                        st.rerun()
                except Exception as e:
                    st.error(f"File padhne mein dikkat: {e}")

        st.markdown("---")
        st.markdown("**Baaki pending jile:**")
        for d in pending_districts:
            if d != pending_sel if 'pending_sel' in dir() else True:
                st.markdown(f"- {d}")
