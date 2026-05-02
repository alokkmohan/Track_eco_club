import pandas as pd
import io
import os
import json
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl

os.makedirs("reports", exist_ok=True)

DISTRICTS = {
    "Lakhimpur Kheri":  ("KHERI",       "KHERI",          True),
    "Auraiya":          ("AURAIYA",      "AURAIYA",        True),
    "Chandauli":        ("CHANDAULI",    "CHANDAULI",      True),
    "Bulandshahr":      ("BULANDSHAHR",  "BULANDSHAHR",    True),
    "Hathras":          ("HATHRAS",      "HATHRAS",        True),
    "Shravasti":        ("SHRAV|SHRAW",  "SHRAV|SHRAW",    False),
    "Gonda":            ("GONDA",        "GONDA",          True),
    "Basti":            ("BASTI",        "BASTI",          True),
    "Bareilly":         ("BAREILLY",     "BAREILLY",       True),
    "Ballia":           ("BALLIA",       "BALLIA",         True),
    "Rampur":           ("RAMPUR",       "RAMPUR",         True),
    "Moradabad":        ("MORADABAD",    "MORADABAD",      True),
    "Etawah":           ("ETAWAH",       "ETAWAH",         True),
    "Azamgarh":         ("AZAMGARH",     "AZAMGARH",       True),
    "Amethi":           ("AMETHI",       "AMETHI",         False),
    "Kannauj":          ("KANNAUJ",      "KANNAUJ",        True),
    "Jaunpur":          ("JAUNPUR",      "JAUNPUR",        True),
    "Hardoi":           ("HARDOI",       "HARDOI",         True),
    "Hapur":            ("HAPUR",        "HAPUR",          False),
    "Amroha":           ("AMROHA",       "AMROHA|JYOTIBA", False),
    "Bhadohi":          ("BHADOHI",      "BHADOI",         False),
    "Prayagraj":        ("PRAYAGRAJ",    "PRAYAGRAJ",      True),
    "Sambhal":          ("SAMBHAL",      "SAMBHAL",        False),
    "Mirzapur":         ("MIRZAPUR",     "MIRZAPUR",       True),
    "Meerut":           ("MEERUT",       "MEERUT",         True),
    "Farrukhabad":      ("FARRUKHABAD",  "FARRUKHABAD",    True),
}

def normalize_udise(val):
    s = str(val).strip().split('.')[0].lstrip('0')
    return s.zfill(10)

def load_secondary_list():
    return {
        sheet: pd.read_excel("data/Final_Secondary_School_List.xlsx", sheet_name=sheet)
        for sheet in ["Govt Schools", "Aided Schools", "Private Schools"]
    }

def load_notifications():
    return pd.read_excel("data/All_Schools_with_Notifications_UP.xlsx")

def get_district_data(notifications_df, secondary_data, dist_label):
    sec_pattern, notif_pattern, exact = DISTRICTS[dist_label]

    if exact:
        notif_d = notifications_df[notifications_df['District'].str.strip().str.upper() == notif_pattern].copy()
    else:
        notif_d = notifications_df[notifications_df['District'].str.strip().str.upper().str.contains(notif_pattern, regex=True, na=False)].copy()

    notif_d['UDISE_norm'] = notif_d['UDISE ID'].apply(normalize_udise)
    notif_set = set(notif_d['UDISE_norm'])

    result = {}
    for sheet, df in secondary_data.items():
        dist_col  = [c for c in df.columns if 'district' in c.lower()][0]
        udise_col = [c for c in df.columns if 'udise' in c.lower()][0]

        if exact:
            d = df[df[dist_col].astype(str).str.strip().str.upper() == sec_pattern].copy()
        else:
            d = df[df[dist_col].astype(str).str.strip().str.upper().str.contains(sec_pattern, regex=True, na=False)].copy()

        d[dist_col]  = dist_label
        d['UDISE_norm'] = d[udise_col].apply(normalize_udise)
        d[udise_col] = d['UDISE_norm']
        d['Upload Status'] = d['UDISE_norm'].apply(lambda x: 'Uploaded' if x in notif_set else 'Pending')
        d = d.drop(columns=['UDISE_norm'])
        result[sheet] = d

    return result

def build_excel(district_data):
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    for sheet, df in district_data.items():
        df_out = df.reset_index(drop=True)
        df_out.index = df_out.index + 1
        df_out.index.name = 'S.No.'
        df_out.to_excel(writer, sheet_name=sheet, index=True)

    writer.close()
    output.seek(0)

    wb = openpyxl.load_workbook(output)
    green_fill  = PatternFill('solid', fgColor='C6EFCE')
    red_fill    = PatternFill('solid', fgColor='FFC7CE')
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))

    for ws in wb.worksheets:
        status_col    = None
        udise_col_idx = None
        for cell in ws[1]:
            if cell.value == 'Upload Status':
                status_col = cell.column
            if cell.value and 'udise' in str(cell.value).lower():
                udise_col_idx = cell.column
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin
                cell.alignment = Alignment(vertical='center')
                if udise_col_idx and cell.column == udise_col_idx:
                    cell.number_format = '@'
                    if cell.value:
                        cell.value = str(cell.value).zfill(10)
            if status_col:
                sc = ws.cell(row=row[0].row, column=status_col)
                if sc.value == 'Uploaded':
                    sc.fill = green_fill
                    sc.font = Font(bold=True, color='276221')
                elif sc.value == 'Pending':
                    sc.fill = red_fill
                    sc.font = Font(bold=True, color='9C0006')
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 45)
        ws.freeze_panes = 'B2'

    final = io.BytesIO()
    wb.save(final)
    final.seek(0)
    return final

def generate_index(summary, generated_time):
    total_grand    = sum(sum(s[sh]["total"]    for sh in ["Govt Schools","Aided Schools","Private Schools"]) for s in summary)
    uploaded_grand = sum(sum(s[sh]["uploaded"] for sh in ["Govt Schools","Aided Schools","Private Schools"]) for s in summary)
    pending_grand  = total_grand - uploaded_grand
    pct_grand      = round(uploaded_grand / total_grand * 100) if total_grand > 0 else 0

    ready   = sorted([s for s in summary if not s.get("priv_missing")], key=lambda x: x["district"])
    pending = sorted([s for s in summary if s.get("priv_missing")],     key=lambda x: x["district"])

    # Build JS data object for ready districts
    js_data_parts = []
    for s in ready:
        dist = s["district"]
        total_d    = sum(s[sh]["total"]    for sh in ["Govt Schools","Aided Schools","Private Schools"])
        uploaded_d = sum(s[sh]["uploaded"] for sh in ["Govt Schools","Aided Schools","Private Schools"])
        pct        = round(uploaded_d / total_d * 100) if total_d > 0 else 0
        js_data_parts.append(
            f'"{dist}":{{'
            f'file:"reports/{dist}_ECO_Club_Status.xlsx",'
            f'total:{total_d},uploaded:{uploaded_d},pending:{total_d-uploaded_d},pct:{pct},'
            f'govt:{{total:{s["Govt Schools"]["total"]},up:{s["Govt Schools"]["uploaded"]},pend:{s["Govt Schools"]["pending"]}}},'
            f'aided:{{total:{s["Aided Schools"]["total"]},up:{s["Aided Schools"]["uploaded"]},pend:{s["Aided Schools"]["pending"]}}},'
            f'pvt:{{total:{s["Private Schools"]["total"]},up:{s["Private Schools"]["uploaded"]},pend:{s["Private Schools"]["pending"]}}}'
            f'}}'
        )
    js_data = "{" + ",".join(js_data_parts) + "}"

    # Dropdown options for ready districts
    ready_options = "\n".join(
        f'      <option value="{s["district"]}">{s["district"]}</option>' for s in ready
    )

    # Pending district dropdown options
    pending_options = "\n".join(
        f'      <option value="{s["district"]}">{s["district"]}</option>' for s in pending
    )

    html = f"""<!DOCTYPE html>
<html lang="hi">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>ECO Club — District Reports</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
  <style>
    body {{ background:#f0f4f8; font-family:'Segoe UI',sans-serif; }}
    .hero {{ background:linear-gradient(135deg,#0d3320,#1a6644); color:#fff; padding:2.5rem 0 2rem; }}
    .stat-box {{ background:rgba(255,255,255,.15); border-radius:10px; padding:.8rem 1.4rem; min-width:110px; }}
    .card {{ border-radius:12px; border:none; }}
    .progress {{ background:#dee2e6; }}
    #distSelect {{ border-radius:10px; font-size:1.05rem; }}
    #distCard {{ transition: all .2s; }}
    .pending-panel {{ border-left:4px solid #ffc107; }}
  </style>
</head>
<body>

<div class="hero">
  <div class="container text-center">
    <h1 class="fw-bold mb-1">🌿 ECO Club</h1>
    <p class="lead mb-2">UP Board Secondary Schools — Notification Upload Status</p>
    <div class="d-flex flex-wrap justify-content-center gap-3 mb-3">
      <div class="stat-box"><div class="fs-4 fw-bold">{total_grand}</div><div class="small">Total Schools</div></div>
      <div class="stat-box" style="background:rgba(40,167,69,.5)"><div class="fs-4 fw-bold">{uploaded_grand}</div><div class="small">Uploaded ✅</div></div>
      <div class="stat-box" style="background:rgba(220,53,69,.5)"><div class="fs-4 fw-bold">{pending_grand}</div><div class="small">Pending ❌</div></div>
      <div class="stat-box"><div class="fs-4 fw-bold">{pct_grand}%</div><div class="small">Completion</div></div>
    </div>
    <div class="small opacity-75">Last updated: {generated_time} &nbsp;|&nbsp; {len(ready)} districts ready &nbsp;|&nbsp; {len(pending)} private list pending</div>
  </div>
</div>

<div class="container py-4">
  <div class="row g-4">

    <!-- LEFT: Ready districts -->
    <div class="col-lg-7">
      <div class="card shadow-sm h-100">
        <div class="card-header fw-bold text-white" style="background:#1F4E79;border-radius:12px 12px 0 0">
          ✅ Report Ready Districts ({len(ready)})
        </div>
        <div class="card-body">
          <label class="form-label fw-semibold mb-1">अपना जिला चुनें</label>
          <select id="distSelect" class="form-select form-select-lg mb-3" onchange="showDist(this.value)">
            <option value="">-- जिला चुनें --</option>
{ready_options}
          </select>

          <div id="distCard" class="d-none">
            <div class="d-flex justify-content-between align-items-center mb-1">
              <span id="dLabel" class="fw-bold fs-5"></span>
              <span id="dPct" class="badge fs-6"></span>
            </div>
            <div class="progress mb-3" style="height:10px;border-radius:5px">
              <div id="dBar" class="progress-bar" style="width:0%"></div>
            </div>
            <table class="table table-sm table-bordered mb-3">
              <thead class="table-light"><tr>
                <th></th><th class="text-center">Total</th>
                <th class="text-center text-success">✅ Uploaded</th>
                <th class="text-center text-danger">❌ Pending</th>
              </tr></thead>
              <tbody>
                <tr><td>🏛️ Govt</td><td id="gT" class="text-center"></td><td id="gU" class="text-center text-success fw-bold"></td><td id="gP" class="text-center text-danger"></td></tr>
                <tr><td>🤝 Aided</td><td id="aT" class="text-center"></td><td id="aU" class="text-center text-success fw-bold"></td><td id="aP" class="text-center text-danger"></td></tr>
                <tr><td>🏫 Private</td><td id="pT" class="text-center"></td><td id="pU" class="text-center text-success fw-bold"></td><td id="pP" class="text-center text-danger"></td></tr>
              </tbody>
            </table>
            <a id="dBtn" href="#" class="btn btn-success w-100 fw-bold" download>
              ⬇️ Report Download करें
            </a>
          </div>
        </div>
      </div>
    </div>

    <!-- RIGHT: Pending districts with client-side upload -->
    <div class="col-lg-5">
      <div class="card shadow-sm h-100 pending-panel">
        <div class="card-header fw-bold bg-warning text-dark" style="border-radius:12px 12px 0 0">
          ⚠️ Private List Pending ({len(pending)})
        </div>
        <div class="card-body d-flex flex-column">
          {'<p class="text-success fw-bold mb-0">सभी जिलों की private list उपलब्ध है! 🎉</p>' if not pending else f"""
          <label class="form-label fw-semibold mb-1">अपना जिला चुनें</label>
          <select id="pendingSelect" class="form-select mb-3" onchange="showPendingDist(this.value)">
            <option value="">-- जिला चुनें --</option>
{pending_options}
          </select>

          <div id="pendingCard" class="d-none flex-grow-1 d-flex flex-column">
            <a id="tmplBtn" href="#" class="btn btn-outline-secondary w-100 mb-3" download>
              📥 Template Download करें
            </a>
            <label class="form-label fw-semibold mb-1">Template भरके यहाँ upload करें:</label>
            <input type="file" id="pvtFile" class="form-control mb-3"
                   accept=".xlsx,.csv" onchange="processPrivateList()">
            <div id="pvtResult" class="d-none mt-auto"></div>
          </div>"""}
        </div>
      </div>
    </div>

  </div>
</div>

<footer class="text-center py-3 text-muted small mt-2">
  ECO Club — UP Board Secondary Schools &nbsp;|&nbsp; Auto-updated on data push
</footer>

<script src="https://cdn.sheetjs.com/xlsx-latest/package/dist/xlsx.full.min.js"></script>
<script>
const DATA = {js_data};

// ── Left panel: Ready district info card ──
function showDist(name) {{
  const card = document.getElementById('distCard');
  if (!name) {{ card.classList.add('d-none'); return; }}
  const d = DATA[name];
  const barColor   = d.pct >= 75 ? 'bg-success' : d.pct >= 40 ? 'bg-warning' : 'bg-danger';
  const badgeColor = d.pct >= 75 ? 'bg-success' : d.pct >= 40 ? 'bg-warning text-dark' : 'bg-danger';
  document.getElementById('dLabel').textContent = '📍 ' + name;
  document.getElementById('dPct').textContent   = d.pct + '%';
  document.getElementById('dPct').className     = 'badge fs-6 ' + badgeColor;
  document.getElementById('dBar').style.width   = d.pct + '%';
  document.getElementById('dBar').className     = 'progress-bar ' + barColor;
  document.getElementById('gT').textContent = d.govt.total;  document.getElementById('gU').textContent = d.govt.up;
  document.getElementById('gP').textContent = d.govt.pend;
  document.getElementById('aT').textContent = d.aided.total; document.getElementById('aU').textContent = d.aided.up;
  document.getElementById('aP').textContent = d.aided.pend;
  document.getElementById('pT').textContent = d.pvt.total;   document.getElementById('pU').textContent = d.pvt.up;
  document.getElementById('pP').textContent = d.pvt.pend;
  document.getElementById('dBtn').href = d.file;
  document.getElementById('dBtn').textContent = '⬇️ ' + name + ' — Report Download करें';
  card.classList.remove('d-none');
}}

// ── Right panel: Pending district upload ──
let _currentPending = '';
let _pvtResult = null;

function showPendingDist(name) {{
  _currentPending = name;
  const card = document.getElementById('pendingCard');
  if (!name) {{ card.classList.add('d-none'); return; }}
  document.getElementById('tmplBtn').href     = `reports/templates/${{name}}_Private_Template.xlsx`;
  document.getElementById('tmplBtn').download = `${{name}}_Private_Template.xlsx`;
  document.getElementById('pvtFile').value    = '';
  document.getElementById('pvtResult').classList.add('d-none');
  card.classList.remove('d-none');
}}

function normalizeUdise(val) {{
  return String(val).split('.')[0].replace(/^0+/, '').padStart(10, '0');
}}

async function processPrivateList() {{
  const file = document.getElementById('pvtFile').files[0];
  if (!file || !_currentPending) return;
  const resultDiv = document.getElementById('pvtResult');
  resultDiv.innerHTML = '<div class="text-muted small text-center py-2">⏳ Processing...</div>';
  resultDiv.classList.remove('d-none');
  try {{
    const resp = await fetch('reports/notif_udise.json?t=' + Date.now());
    const notifData = await resp.json();
    const uploadedSet = new Set(notifData[_currentPending] || []);

    const buf  = await file.arrayBuffer();
    const wb   = XLSX.read(buf);
    const ws   = wb.Sheets[wb.SheetNames[0]];
    const rows = XLSX.utils.sheet_to_json(ws, {{defval: ''}});

    if (!rows.length) {{
      resultDiv.innerHTML = '<div class="alert alert-warning">File mein koi data nahi mila।</div>';
      return;
    }}
    const udiseKey = Object.keys(rows[0]).find(k => k.toLowerCase().includes('udise'));
    if (!udiseKey) {{
      resultDiv.innerHTML = '<div class="alert alert-danger">❌ UDISE column nahi mila। Template use karein।</div>';
      return;
    }}
    _pvtResult = rows.map(row => {{
      const udise  = normalizeUdise(String(row[udiseKey]));
      const status = uploadedSet.has(udise) ? 'Uploaded' : 'Pending';
      const out    = {{}};
      for (const k of Object.keys(row)) out[k] = row[k];
      out[udiseKey]       = udise;
      out['Upload Status'] = status;
      return out;
    }});
    const upCt  = _pvtResult.filter(r => r['Upload Status'] === 'Uploaded').length;
    const pnCt  = _pvtResult.length - upCt;
    const pct   = Math.round(upCt / _pvtResult.length * 100);
    const barcl = pct >= 75 ? 'bg-success' : pct >= 40 ? 'bg-warning' : 'bg-danger';
    resultDiv.innerHTML = `
      <div class="fw-bold mb-2">🏫 ${{_currentPending}} — Private Schools</div>
      <div class="d-flex justify-content-between mb-1">
        <small>${{upCt}} / ${{_pvtResult.length}} uploaded</small>
        <small class="fw-bold">${{pct}}%</small>
      </div>
      <div class="progress mb-3" style="height:8px">
        <div class="${{barcl}} progress-bar" style="width:${{pct}}%"></div>
      </div>
      <div class="row g-2 mb-3 text-center small">
        <div class="col-4"><div class="bg-light rounded p-2"><b>${{_pvtResult.length}}</b><br>Total</div></div>
        <div class="col-4"><div class="bg-success bg-opacity-25 rounded p-2 text-success"><b>${{upCt}}</b><br>✅ Uploaded</div></div>
        <div class="col-4"><div class="bg-danger bg-opacity-25 rounded p-2 text-danger"><b>${{pnCt}}</b><br>❌ Pending</div></div>
      </div>
      <button onclick="downloadPvtReport()" class="btn btn-success w-100 fw-bold">⬇️ Report Download करें</button>`;
  }} catch(e) {{
    resultDiv.innerHTML = `<div class="alert alert-danger">Error: ${{e.message}}</div>`;
  }}
}}

function downloadPvtReport() {{
  if (!_pvtResult) return;
  const wb = XLSX.utils.book_new();
  const ws = XLSX.utils.json_to_sheet(_pvtResult);
  const range = XLSX.utils.decode_range(ws['!ref'] || 'A1');
  let statusCol = -1;
  for (let c = range.s.c; c <= range.e.c; c++) {{
    const addr = XLSX.utils.encode_cell({{r:0, c}});
    if (ws[addr] && ws[addr].v === 'Upload Status') {{ statusCol = c; break; }}
  }}
  if (!ws['!cols']) ws['!cols'] = [];
  for (let c = range.s.c; c <= range.e.c; c++) ws['!cols'][c] = {{wch: 18}};
  XLSX.utils.book_append_sheet(wb, ws, 'Private Schools');
  XLSX.writeFile(wb, `${{_currentPending}}_Private_Status.xlsx`);
}}
</script>
</body>
</html>"""

    with open("index.html", "w", encoding="utf-8") as f:
        f.write(html)


def make_template(dist_label):
    df = pd.DataFrame(columns=['District Name', 'Block Name', 'School Name', 'UDISE Code', 'Board', 'Managed By'])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Private Schools')
    output.seek(0)
    return output.read()


def main():
    print("Loading data files...")
    notifications_df = load_notifications()
    secondary_data   = load_secondary_list()

    os.makedirs("reports/templates", exist_ok=True)

    # Build UDISE lookup per district for client-side JS matching
    notif_udise = {}
    for dist_label, (_, notif_pattern, exact) in DISTRICTS.items():
        if exact:
            nd = notifications_df[notifications_df['District'].str.strip().str.upper() == notif_pattern]
        else:
            nd = notifications_df[notifications_df['District'].str.strip().str.upper().str.contains(notif_pattern, regex=True, na=False)]
        notif_udise[dist_label] = [normalize_udise(u) for u in nd['UDISE ID'].dropna().tolist()]
    with open("reports/notif_udise.json", "w") as f:
        json.dump(notif_udise, f)
    print("  notif_udise.json written")

    summary = []
    for dist_label in sorted(DISTRICTS.keys()):
        print(f"  Generating: {dist_label}...")
        district_data = get_district_data(notifications_df, secondary_data, dist_label)
        excel_bytes   = build_excel(district_data)

        with open(f"reports/{dist_label}_ECO_Club_Status.xlsx", "wb") as f:
            f.write(excel_bytes.read())

        stats = {"district": dist_label}
        for sheet in ["Govt Schools", "Aided Schools", "Private Schools"]:
            df = district_data[sheet]
            total    = len(df)
            uploaded = int((df['Upload Status'] == 'Uploaded').sum())
            stats[sheet] = {"total": total, "uploaded": uploaded, "pending": total - uploaded}

        stats["priv_missing"] = (stats["Private Schools"]["total"] == 0)
        if stats["priv_missing"]:
            tmpl_bytes = make_template(dist_label)
            with open(f"reports/templates/{dist_label}_Private_Template.xlsx", "wb") as f:
                f.write(tmpl_bytes)
            print(f"    -> template saved (private list missing)")

        summary.append(stats)

    generated_time = datetime.utcnow().strftime("%d %b %Y, %H:%M UTC")
    generate_index(summary, generated_time)
    print(f"\nDone! {len(DISTRICTS)} districts | index.html + notif_udise.json generated.")

if __name__ == "__main__":
    main()
